import requests
import os
import zipfile
import subprocess
from collections import defaultdict
import time
from datetime import datetime
from adbutils import adb
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import pandas as pd
import git
import sys
import re
from bs4 import BeautifulSoup
import argparse

threadhold_cpu_max = 150
threadhold_cpu_mean = 10
threadhold_mem_max = 250
threadhold_mem_mean = 200

remove_forward_all = "adb forward --remove-all"
remove_reverse_all = "adb reverse --remove-all"
REVERSE_DEVICE_PORT = 3333
REVERSE_PC_PORT = 5555
FORWARD_PORT = 4444
retry_times = 0

def extract_zip_path_info(jenkins_pack_job_name, brand):
    # Jenkins Pack Job Name:
    jenkins_pack_job_name_list = {
        'DriveCube_app_48_debug' : '_hcp3_48',
        'DriveCube_app_48_release' : '_hcp3_48',
        'DriverCube_ut_report' : '_hcp3_48',
        'SmartSystem_app' : '_hcp3_46',
        'SmartSystem_app_48_debug' : '_hcp3_48',
        'SmartSystem_app_48_release' : '_hcp3_48',
        'SmartSystem_app_55_release' : '_oi_55',
        'SmartSystem_ut_report' : '_hcp3_46',
        'DriveCube_app' : '_hcp3_48'
    }
    jfrog_domain = 'jfrog.maezia.com'
    for target_zip_keyword, target_release_folder_in_zip in jenkins_pack_job_name_list.items():
        if jenkins_pack_job_name == target_zip_keyword:
            return target_zip_keyword, f'{brand}{target_release_folder_in_zip}' , jfrog_domain
    print("=====WRONG jenkins_pack_job_name====")
    exit()

def writeFile(filename, mode, content):
    with open(filename, mode=mode, encoding='utf-8') as f:
        f.write(content + "\n")

def readFile(filename, mode):
    with open(filename, 'r', encoding='utf-8') as f:
        content = f.read()
        return content
    
def deleteFile(file_path):
    if os.path.exists(file_path):
        os.remove(file_path)
        print(f"{file_path} 已删除。")
    else:
        print(f"{file_path} 不存在。")

def rename(old_name, new_name):
    if os.path.exists(old_name):
        os.rename(old_name, new_name)
        print(f"文件已重命名为 {new_name}。")
    else:
        print(f"{old_name} 文件不存在。")

def extract_name_and_extension(file_path):
    base_name = os.path.basename(file_path)
    file_name, file_extension = os.path.splitext(base_name)
    return file_name, file_extension

def mkdirs(directory):
    try:
        os.makedirs(directory, exist_ok=True)
        print("Directory '%s' created successfully" % directory)
    except Exception as e:
        print("An error occurred: %s" % str(e))

def login():    
    global retry_times
    # 登录URL，需要替换成实际的登录接口
    login_url = 'https://jfrog-prod.maezia.com/ui/api/v1/ui/auth/login?_spring_security_remember_me=true'                 

    # 登录所需的参数，需要根据实际登录表单进行调整
    login_data = {
        'user': 'usbrgwu',
        'password': 'Cariad1234567!',
        'type' : 'login'
    }

    headers = {        
        'x-requested-with':'XMLHttpRequest'
    }
    # 创建会话
    session = requests.Session()
    
    # 发送POST请求    
    response = session.post(login_url, headers=headers, data=login_data)    
    print(response)
    
    # 检查响应状态
    if response.status_code == 200:
        print("登录成功！")
        return session
    else:                                
        if(retry_times < 3):
            retry_times += 1
            print(f"登录失败，准备重试{retry_times}/3")
            login()
        else:
            return False

def getSmartSystemZips(session, url):
    # 使用requests库发送GET请求，并传递headers字典
    response = session.get(url)    
    # 检查响应状态
    if response.status_code == 200:
        print("Smartsystem release包列表获取成功！")
        return response.text
    else:
        print("Smartsystem release包列表获取失败，请检查网络")
        return False

"""
def downloadZip(session, url , zip_folder, file_name):    
    zip_filename = zip_folder + '/' + file_name
    chunk_size = 8192 
    response = session.get(url, stream=True)
    if response.status_code == 200:        
        mkdirs(zip_folder)        
        print('正在下载' + zip_filename)
        with open(zip_filename, 'wb') as file:
            i = 0
            # 逐块读取响应内容并写入文件
            for chunk in response.iter_content(chunk_size=chunk_size):
                i += 1
                if chunk:  # 过滤掉keep-alive新块
                    file.write(chunk)
                if i % 1000 == 0:
                    print('下载到第' + str(i) + '个数据包了,chunk_zie=' + str(chunk_size))
            print(zip_filename + ':下载完毕')
    else:
        print('Failed to download file')
        exit()
    return zip_filename
"""

def downloadZip(session, url, zip_folder, file_name):
        zip_filename = zip_folder + "/" + file_name
        if os.path.exists(zip_filename):
            print("之前已经下载成功，跳过本步骤")
            return zip_filename

        chunk_size = 8192
        print(url)
        exit()
        response = session.get(url, stream=True)
        if response.status_code == 200:
            mkdirs(zip_folder)
            print("正在下载" + zip_filename)
            with open(zip_filename, "wb") as file:
                i = 0
                # 逐块读取响应内容并写入文件
                for chunk in response.iter_content(chunk_size=chunk_size):
                    i += 1
                    if chunk:  # 过滤掉keep-alive新块
                        file.write(chunk)
                    if i % 1000 == 0:
                        print("下载到第" + str(i) + "个数据包了,chunk_zie=" + str(chunk_size))
                print(zip_filename + ":下载完毕")
        else:
            print("Failed to download file")
            exit()
        return zip_filename


def unzip(zip_path, extract_path):
    unzipped_files = defaultdict(lambda: defaultdict(str)) 
    file_name, _ = extract_name_and_extension(zip_path)
    unzip_folder = extract_path + '/' + file_name
    mkdirs(unzip_folder)

    # 打开ZIP文件
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        # 解压所有文件
        zip_ref.extractall(unzip_folder)

        # 列出所有文件
        for file in zip_ref.namelist():
            parts = file.split('/')            
            if len(parts) < 2:
                continue
            folder = parts[0]
            name = parts[1]
            filename, extension_name = extract_name_and_extension(name)
            name = unzip_folder + '/' + folder + '/' + name
            #print(name, filename, extension_name)
            if extension_name == '.apk':
                unzipped_files[folder]['apk'] = name
            elif extension_name == '.xml' and filename.find('default_permissions'):
                unzipped_files[folder]['default_permissions'] = name
            elif extension_name == '.xml' and filename.find('privapp_permissions'):
                unzipped_files[folder]['privapp_permissions'] = name
        print('解压缩成功！')
        return unzipped_files
        

def getTheLatestHcp3ReleaseZip(session, zip_folder, sorted_zips, jfrog_dl_base_url, filter="hcp3_release"):    
    for row in sorted_zips:
        print(row['name'])
        if filter in row['name']:            
            zip_url = jfrog_dl_base_url.format(row['name'])            
            return downloadZip(session, zip_url, zip_folder, row['name'])            
    return False

def runShell(command):
    print('正在执行命令：' + command)
    result = subprocess.run(command, shell=True, stdout=subprocess.PIPE)
    str = result.stdout.decode('utf-8')    
    return str


def startCaptureTopLogs(testbench_ip, log_path):
    today = datetime.now().strftime('%y%m%d')
    # 构建日志文件路径
    mkdirs(log_path + f"/{testbench_ip}")
    log_file_path = log_path + f"/{testbench_ip}/{today}.log"

    # 执行top命令并将输出重定向到日志文件
    result = runShell(f'adb connect {testbench_ip}')
    print(result)
    result = runShell(f'adb root') 
    print(result)
    result = runShell(f'adb remount')
    print(result)
    result = runShell(f"adb shell \"top -b|grep -E 'technology.cariad.smartsystem.cn|%cpu'\" > {log_file_path} &")
    print(result)
    return log_file_path

def readlines(filename, max_lines):
    lines = []
    with open(filename, 'r', encoding='utf-8') as file:
        for i, line in enumerate(file):
            if i == max_lines:
                break
            lines.append(line)
    return lines

def getTopPid(line):
    print(line)
    columns = line.split(' ')
    return columns[0]

def forward():
    try:
        # adb forward --remove-all
        runShell(remove_forward_all)

        # adb forward  tcp:14714 tcp:14714
        result_forward = runShell(f"adb forward tcp:{FORWARD_PORT} tcp:{FORWARD_PORT}")
        if result_forward:
            commandResult = result_forward.splitlines()[0]
            read = commandResult.replace(" ", "")
            if read.__contains__("\n"):
                read = read.replace("\n", "")
            # log("forward:" + str(line))
            time.sleep(1)
            return read
        else:
            print("forward Port already binding!")
    except Exception as e_forward:
        print(f"forward command error：{e_forward}")


def reverse():
    try:
        # adb reverse --remove-all
        runShell(remove_reverse_all)
        # adb reverse tcp:5554 tcp:4444
        result_reverse = runShell(
            f"adb reverse tcp:{REVERSE_DEVICE_PORT} tcp:{REVERSE_PC_PORT}"
        )
        if result_reverse:
            commandResult = result_reverse.splitlines()[0]
            if commandResult.__contains__("\n"):
                commandResult = commandResult.replace("\n", "")
                time.sleep(1)
            return commandResult
    except Exception as e_reverse:
        print(f'"reverse command error:{e_reverse}"')


from datetime import datetime
from adbutils import adb

def executeAdbTopCommand(testbench_ip, command):    
    client = adb.device(testbench_ip)
    result = client.shell(command)
    print(result)


def copySheet(source_file, target_file, source_sheet_name, target_sheet_name=None):
    # 加载源工作簿和目标工作簿
    source_workbook = load_workbook(source_file)
    target_workbook = load_workbook(target_file)

    # 获取源工作表
    source_sheet = source_workbook[source_sheet_name]

    # 如果未指定目标工作表名称，则使用源工作表名称
    if target_sheet_name is None:
        target_sheet_name = source_sheet_name

    # 创建目标工作表（如果不存在）
    if target_sheet_name not in target_workbook.sheetnames:
        target_workbook.create_sheet(title=target_sheet_name)
    target_sheet = target_workbook[target_sheet_name]

    # 复制单元格数据
    for row in source_sheet.iter_rows(values_only=True):
        target_sheet.append(row)

    # 保存目标工作簿
    target_workbook.save(target_file)


def getStatistic(filename, sheetname, column_name):    
    import pandas as pd
     # 读取Excel数据
    if sheetname =='cpuinfo':
        scale = 1        
        unit = '%'        
        df = pd.read_excel(filename , header=[1], sheet_name=sheetname, engine='openpyxl')
    else:
        scale = 1024
        unit = 'MB'
        df = pd.read_excel(filename , sheet_name=sheetname, engine='openpyxl')    
    mean_value = round(df[column_name].mean() / scale, 1)
    max_value = round(df[column_name].max() / scale, 1)

    return mean_value, max_value, unit

def getApkVersion(package_name, testbench_ip = ''):
    versionName = ''
    SystemVersion = ''    
    result = runShell(f'adb shell dumpsys package {package_name} | findstr "versionName"')
    print(result)
    temp = result.split("\n")
    try:
        if len(temp) > 1:
            versionName = temp[0].strip().split("=")[1]
            if "=" in temp[1]:
                SystemVersion = temp[1].strip().split("=")[1]
            else:
                SystemVersion = versionName
        else:
            versionName = temp[0].strip().split("=")[1]
            SystemVersion = temp[0].strip().split("=")[1]
    except Exception as e:
        # 处理其他类型的异常
        print(f"发生了其他错误：{e}")
    return versionName, SystemVersion


def generateFinalReport(excel_file, pack_name, begin_time, end_time):
    # 打开 Excel 文件
    workbook = load_workbook('./SmartSystemPerformanceTest.xlsx')

    # 选择工作表
    sheet = workbook.active

    versionName, SystemVersion = getApkVersion(pack_name)
    # 填写数据
    sheet['B5'] = SystemVersion
    sheet['B6'] = versionName

    # 设置背景色为红色
    color_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    color_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # CPU的值
    cpu_mean, cpu_max , unit = getStatistic(excel_file, 'cpuinfo', 'Total')
    testresult_cpumax = cpu_max < threadhold_cpu_max
    testresult_cpumean = cpu_mean < threadhold_cpu_mean
    sheet['D9'] = f'{cpu_max}{unit}'
    sheet['D10'] = f'{cpu_mean}{unit}'
    sheet['G9'] = 'PASS' if testresult_cpumax else 'FAILED'
    sheet['G10']  = 'PASS' if testresult_cpumean else 'FAILED'
    sheet['G9'].fill = color_green if testresult_cpumax else color_red
    sheet['G10'].fill  = color_green if testresult_cpumean else color_red

    # 内存的值
    mem_mean, mem_max , unit = getStatistic(excel_file, 'meminfo', 'Total')
    testresult_memmax = mem_max < threadhold_mem_max
    testresult_memmean = mem_mean < threadhold_mem_mean
    sheet['D11'] = f'{mem_max}{unit}'
    sheet['D12'] = f'{mem_mean}{unit}'
    sheet['G11'] = 'PASS' if testresult_memmax else 'FAILED'
    sheet['G12']  = 'PASS' if testresult_memmean else 'FAILED'
    sheet['G11'].fill = color_green if testresult_memmax else color_red
    sheet['G12'].fill = color_green if testresult_memmean else color_red
    
    #分析crash/OOM的日志，并填写报告：
    crash_num, oom_num, crash_logs, oom_logs = getFinalCrashOOMTestResult()
    sheet['C13'] = 'Total Times'
    sheet['C14'] = 'Total Times'
    sheet['D13'] = crash_num
    sheet['D14'] = oom_num
    sheet['G13'] = 'PASS' if crash_num == 0 else 'FAILED'
    sheet['G14'] = 'PASS' if oom_num == 0 else 'FAILED'
    sheet['G13'].fill = color_green if crash_num == 0 else color_red
    sheet['G14'].fill = color_green if oom_num == 0 else color_red
    sheet['H13'] = crash_logs
    sheet['H14'] = oom_logs

    # 测试开始时间，测试持续时长：
    sheet['G17'] = time.strftime("%Y-%m-%d %H:%M", time.localtime(begin_time))
    sheet['H17'] = getTimeInterval(begin_time, end_time)    

    # 测试最终结果：    
    final_result = getFinalTestResult(testresult_cpumax, testresult_cpumean, testresult_memmax, testresult_memmean, crash_num, oom_num)
    sheet['C4'] = final_result
    sheet['G4'].fill = color_green if final_result == 'PASS' else color_red

    #写入测试日志：
    test_log = readFile('temp.txt', 'r')
    test_log.replace('\n', '\r\n')
    sheet['C18'] = test_log

    #C15:
    description = sheet['C15'].value
    description.replace('\n', '\r\n')
    sheet['C15'] = description

    # 保存文件
    workbook.save('./SmartSystemPerformanceTest_updated.xlsx')

    # 合并报表
    copySheet(excel_file, './SmartSystemPerformanceTest_updated.xlsx', 'cpuinfo', 'cpuinfo')
    copySheet(excel_file, './SmartSystemPerformanceTest_updated.xlsx', 'meminfo', 'meminfo')

    #删除临时文件：
    deleteFile('temp.txt')
    deleteFile(excel_file)
    rename('./SmartSystemPerformanceTest_updated.xlsx', excel_file)


def getFinalCrashOOMTestResult():    
    #Crash的次数：    
    crash_result = []   
    crash_num = 0 
    for item in os.listdir('result/tmp/crash'):
        if os.path.isfile('result/tmp/crash/' + item):
            tmp = item.split('.crash.')
            if len(tmp) > 1:
                version = tmp[0]
                print(tmp)
                parts = tmp[1].split('.')
                crash_times = int(parts[0])
                crash_result.append(version + ': ' + str(crash_times))
                crash_num += crash_times    
    crash_logs = "\r\n".join(crash_result)
    #oom的次数：
    oom_num = 0
    oom_result = []
    for item in os.listdir('result/tmp/dump'):
        if os.path.isfile('result/tmp/dump/' + item):
            tmp = item.split('.')
            oom_times = int(tmp[1])            
            oom_result.append(item + ': ' + str(oom_times))
            oom_num += oom_times
    oom_logs = "\r\n".join(oom_result)
    return crash_num, oom_num, crash_logs, oom_logs

def getFinalTestResult(testresult_cpumax, testresult_cpumean, testresult_memmax, testresult_memmean, crash_num, oom_num):
    result = []
    if testresult_cpumax == False:
        result.append('CPU usage peak value failed')
    if testresult_cpumean == False:
        result.append('CPU usage average value failed')        
    if testresult_memmax == False:        
        result.append('Memory usage peak value failed')
    if testresult_memmean == False:        
        result.append('Memory usage average value failed')
    if crash_num > 0:        
        result.append('Crash times failed')
    if oom_num > 0:        
        result.append('OOM failed')
    if len(result) == 0:
        result.append('PASS')
    return ','.join(result)
    
def getHtmlBody(excel_file_path, sheet_name):
    df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
    df.fillna('')
    html_table = df.to_html(index=False)    
    html_table = html_table.replace('NaN', '')
    return html_table

def getTimeInterval(begin_time, end_time):
    time_diff = end_time - begin_time
    days, remainder = divmod(time_diff, 86400)
    hours, remainder = divmod(remainder, 3600)
    minutes, seconds = divmod(remainder, 60)    
    human_readable_time = f"{int(hours)} hours, {int(minutes)} minutes"
    return human_readable_time

def getCrashFileFromTestBench(processname):
    mkdirs('result/tmp/crash')
    versionName, SystemVersion = getApkVersion(processname)    
    current_user = runShell('adb shell am get-current-user').strip()
    result = runShell(f'adb root')
    result = runShell(f'adb shell \"ls -l /data/user/{current_user}/{processname}/cache/crash | grep "^-" | wc -l\"')
    if 'No such file or directory' in result:
        total_crash = 0
    else:
        total_crash = int(result)
    zip_filename = f'{versionName}.crash.{total_crash}.tar.gz'
    result = runShell(f'adb shell tar -czvf /data/user/{current_user}/{processname}/cache/{zip_filename} /data/user/{current_user}/{processname}/cache/crash')
    result = runShell(f'adb pull /data/user/{current_user}/{processname}/cache/{zip_filename} result/tmp/crash/{zip_filename}')
    result = runShell(f'adb shell rm -rf /data/user/{current_user}/{processname}/cache/{zip_filename}')
    result = runShell(f'adb shell rm -rf /data/user/{current_user}/{processname}/cache/crash')

def getOOMFileFromTestBench(processname):
    oom_folder = 'dump'    
    mkdirs(f'result/tmp/{oom_folder}')
    current_user = runShell('adb shell am get-current-user').strip()
    result = runShell(f'adb root')
    result = runShell(f'adb shell ls -l /data/user/{current_user}/{processname}/cache/{oom_folder} | grep "^-" | wc -l')
    if 'No such file or directory' in result:
        total_dumps = 0
    else:
        total_dumps = int(result)
    oom_file = f'outOfMemort.{total_dumps}.hprof.tar.gz'        
    result = runShell(f'adb shell tar -czvf /data/user/{current_user}/{processname}/cache/{oom_file} /data/user/{current_user}/{processname}/cache/{oom_folder}')
    result = runShell(f'adb pull /data/user/{current_user}/{processname}/cache/{oom_file} result/tmp/{oom_folder}/{oom_file}')
    result = runShell(f'adb shell rm -rf /data/user/{current_user}/{processname}/cache/{oom_file}')
    result = runShell(f'adb shell rm -rf /data/user/{current_user}/{processname}/cache/{oom_folder}')    

def getLatestGitTag(repo_base, keyword):
    repo = git.Repo(repo_base)
    # 获取所有的标签
    all_tags = [tag.name for tag in repo.tags]
    # 筛选包含特定关键字的标签    
    matching_tags = [tag for tag in all_tags if keyword in tag]
    # 对匹配的标签按版本号进行排序（假设标签格式类似 v1.0、v2.0 等）
    sorted_matching_tags = sorted(matching_tags, key=lambda x: [int(part) if part.isdigit() else part for part in x.split('.')])
    # 获取最近的两个标签
    recent_two_tags = sorted_matching_tags[-2:]
    return recent_two_tags

def getParameterForPython(param_index):
    if param_index < len(sys.argv):
        return sys.argv[param_index]
    else:
        return ''

def getGitLog(repo_base, tag1_name, tag2_name):    
    kpm_tickets = []
    jira_tickets = []
    repo = git.Repo(repo_base)
    tag1 = repo.tags[tag1_name]
    tag2 = repo.tags[tag2_name]
    commits = list(repo.iter_commits(f'{tag1}..{tag2}'))
    for commit in commits:        
        if isKpm(commit.message):
            kpm_tickets.append(commit)                        
        else:
            jira_tickets.append(commit)
    return kpm_tickets, jira_tickets

def isKpm(string):    
    pattern = r'.*(10\d{6}|9\d{6}).*'
    return re.match(pattern, string)

def isJiraTicket(string):
    pattern = r'.*(d{6}).*'
    return re.match(pattern, string)

def sendEmail(mail, title, content):
    msg_to = ['yiwukun@sina.com', 
              'wukun.yi@cariad-technology.cn', 
              'xinhui.ren@cariad-technology.cn',
              'jinming.gao@cariad-technology.cn',
              'zengfu.fan@cariad-technology.cn',              
              'extern.yang.liu1@cariad-technology.cn' , 
              'extern.yunfei.ma@cariad-technology.cn',
              'extern.chunlei.wang@cariad-technology.cn', 
              'extern.jianjun.xi@cariad-technology.cn',
              'extern.xuan.luo1@cariad-technology.cn',
              'extern.chongyang.gao1@cariad-technology.cn',              
              'allenxia@jidouauto.com',
              'hedygao@jidouauto.com',
              'jaxzhang@jidouauto.com',
              'pippohan@jidouauto.com',
              'amyliu@jidouauto.com',
              'bobohu@jidouauto.com',
              'extern.wenyan.ma1@cariad-technology.cn',
              'extern.pin.chai1@cariad-technology.cn',
              'extern.feifei.lv@cariad-technology.cn',
              'extern.mengyao.wu@cariad-technology.cn'
            ]        
    #msg_to = ['wukun.yi@cariad-technology.cn']
    mail.send(msg_to , title , content)

def getHtmlFromExcel(excel_file):
    import xlsx2html
    html_file = "test.html"
    xlsx2html.xlsx2html(excel_file, html_file)
    html = readFile(html_file, 'r').strip()
    html = confirmEnter2Br(html)    
    html = html.replace('[CL]', '<br/>')
    #print(html)
    #writeFile('test_new.html','w',html)
    #exit()
    deleteFile(html_file)
    return html

def confirmEnter2Br(html):
    soup = BeautifulSoup(html, 'html.parser')
    for td in soup.find_all('td'):
        td.string = td.string.replace('\n', '[CL]') if td.string else td.string    
    return str(soup)